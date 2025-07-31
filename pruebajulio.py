from dotenv import load_dotenv
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine
import pandas as pd
from datetime import datetime
import os
load_dotenv()  # Carga las variables del archivo .env

# 📁 Archivo de clasificación
df_file = "Clasificación de funcionarios por Unidad de negocio.xlsx"

# 📅 Determinar hoja del mes anterior
es_meses = {
    'January':'Enero','February':'Febrero','March':'Marzo',
    'April':'Abril','May':'Mayo','June':'Junio','July':'Julio',
    'August':'Agosto','September':'Septiembre',
    'October':'Octubre','November':'Noviembre','December':'Diciembre'
}
now = datetime.now()
sheet_name = f"{es_meses[now.strftime('%B')]} {now.year}".upper()

# 📊 Leer datos y copia para resumen completo
df = pd.read_excel(df_file, sheet_name=sheet_name, dtype=str)
df.columns = df.columns.str.strip()
df_original = df.drop_duplicates(subset=["CEDULA"])

# 🔄 Mapear casos específicos
mask_cali = (df['ÁREA']=='DIRECCION REGIONAL BETA CALI') & (df['UNIDAD DE NEGOCIO'].str.upper()=='SALVADOR')
df.loc[mask_cali,'UNIDAD DE NEGOCIO'] = 'OTROS PROYECTOS'
mask_nac = df['ÁREA']=='DIRECCION NACIONAL DE COBRANZAS BETA'
df.loc[mask_nac,'UNIDAD DE NEGOCIO'] = (
    df.loc[mask_nac,'UNIDAD DE NEGOCIO']
      .str.replace(r'UGP\s*-\s*','',regex=True)
      .str.strip()
)

# 🔗 Unificar CARTERA
df['UNIDAD DE NEGOCIO'] = df['UNIDAD DE NEGOCIO'].apply(
    lambda x: 'CARTERA UNIDAD DE GESTION PERSONALIZADA' if 'CARTERA' in str(x).upper() else x
)
# 🚫 Filtrar CALL CENTER
df = df[~df['ÁREA'].str.contains('CALL', case=False, na=False) & 
         ~df['UNIDAD DE NEGOCIO'].str.contains('CALL', case=False, na=False)]

# 📖 Diccionarios
general_units = [
    {'UN':'100','CCOS':'110101','UNIDAD':'VIGENTE'},
    {'UN':'200','CCOS':'220101','UNIDAD':'JURIDICO HIPOTECARIO'},
    {'UN':'200','CCOS':'220102','UNIDAD':'JURIDICO CONSUMO'},
    {'UN':'300','CCOS':'330101','UNIDAD':'CASTIGADA'},
    {'UN':'500','CCOS':'550101','UNIDAD':'PROPIA'},
    {'UN':'700','CCOS':'770101','UNIDAD':'OTROS PROYECTOS'},
    {'UN':'600','CCOS':'660104','UNIDAD':'CARTERA UNIDAD DE GESTION PERSONALIZADA'}
]
admin_areas = {
    'AREA DE CUMPLIMIENTO BETA': {'UN':'999','CCOS':'990103'},
    'AUDITORIA BETA': {'UN':'999','CCOS':'990102'},
    'DEPTO. GESTION FINANCIERA Y CUMPLIMIENTO BETA': {'UN':'999','CCOS':'990202'},
    'DIRECCION ADMINISTRATIVA Y FINANCIERA BETA': {'UN':'999','CCOS':'990201'},
    'DIRECCION DE RECURSOS HUMANOS': {'UN':'999','CCOS':'990301'},
    'DIRECCION DE SISTEMAS BETA': {'UN':'999','CCOS':'990203'},
    'GERENCIA GENERAL PROMOCIONES Y COBRANZAS BETA': {'UN':'999','CCOS':'990101'},
    'DIRECCION JURIDICA BETA SINCELEJO': {'UN':'999','CCOS':'990103'},
    'DEPTO. DE OPERACION Y ADMINISTRATIVO':{'UN':'999', 'CCOS':'990204'}

}
co_map = {**{area:'001' for area in admin_areas},
    **{f"DIRECCION REGIONAL BETA {c}": code for c,code in [
        ('ARMENIA','002'),('BARRANQUILLA','003'),('BUCARAMANGA','004'),
        ('CALI','005'),('CARTAGENA','007'),('CUCUTA','008'),
        ('IBAGUE','009'),('MANIZALES','010'),('MEDELLIN','011'),
        ('SINCELEJO','012'),('MONTERIA','012'),('NEIVA','013'),
        ('PEREIRA','015'),('SANTA MARTA','016'),('VALLEDUPAR','018'),
        ('TUNJA','019'),('VILLAVICENCIO','019')
    ]},
    'DIRECCION DIGITAL Y DE NEGOCIOS ESPECIALES BETA': '001',
    'DIRECCION NACIONAL DE COBRANZAS BETA': '001',
    'DIRECCION UNIDAD DE VISITAS BETA': '001',
    'JEFATURA JURIDICA BETA': '001'
}

# 🧮 Construir filas y ajustar conteo
general_ugp = df_original['CEDULA'][df_original['UNIDAD DE NEGOCIO'].str.contains('UGP - JURIDICO HIPOTECARIO', case=False, na=False)].nunique()
rows = []
for area in sorted(df['ÁREA'].unique()):
    if area in admin_areas:
        cnt = df[df['ÁREA']==area].shape[0]
        rows.append({
            'CONCAT': area + 'ADMINISTRATIVA', 'ÁREA': area,
            'UNIDAD DE NEGOCIO': 'ADMINISTRATIVA', 'CO': co_map.get(area,''),
            'UN': admin_areas[area]['UN'], 'CCOS': admin_areas[area]['CCOS'], 'Cant': cnt
        })
    else:
        for u in general_units:
            base = df[(df['ÁREA']==area) & (df['UNIDAD DE NEGOCIO']==u['UNIDAD'])].shape[0]
            if area=='DIRECCION NACIONAL DE COBRANZAS BETA' and u['UNIDAD']=='JURIDICO HIPOTECARIO':
                cnt = base - general_ugp
            else:
                cnt = base
            rows.append({
                'CONCAT': area + u['UNIDAD'], 'ÁREA': area, 'UNIDAD DE NEGOCIO': u['UNIDAD'],
                'CO': co_map.get(area,''), 'UN': u['UN'], 'CCOS': u['CCOS'], 'Cant': cnt
            })
# ⚙️ Agregar fila UGP específica
if general_ugp > 0:
    rows.append({
        'CONCAT': 'DIRECCION NACIONAL DE COBRANZAS BETAUGP - JURIDICO HIPOTECARIO',
        'ÁREA': 'DIRECCION NACIONAL DE COBRANZAS BETA', 'UNIDAD DE NEGOCIO': 'UGP - JURIDICO HIPOTECARIO',
        'CO': '020', 'UN': '200', 'CCOS': '220101', 'Cant': general_ugp
    })

# 🖥 DataFrame final y Categoría
res_df = pd.DataFrame(rows)
res_df.loc[res_df['ÁREA']=='DIRECCIÓN REGIONAL BETA TUNJA','CO'] = '019'

# 📝 Resumen con todos los funcionarios
df_sum = df_original.groupby(['ÁREA','UNIDAD DE NEGOCIO'], as_index=False).agg(Cantidad=('CEDULA','count'))

# 💾 Guardar y formatear Excel
with pd.ExcelWriter('Resultado_distribucion.xlsx', engine='openpyxl') as writer:
    res_df.to_excel(writer, index=False, sheet_name='Resultado')
    df_sum.to_excel(writer, index=False, sheet_name='Resumen')
    ws = writer.book['Resultado']
    # Encabezados H y I
    ws.cell(row=1, column=8, value='% Total')
   
    # Suma en G150 y fondo azul
    ws['G150'] = '=SUM(G2:G149)'
    ws['G150'].fill = PatternFill(fill_type='solid', fgColor='0000FF')
    # Porcentajes en H2:H149
    for r in range(2, 150):
        ws.cell(row=r, column=8, value=f'=G{r}/$G$150').number_format = '0.00%'
    # Suma porcentaje H150
    ws['H150'] = '=SUM(H2:H149)'
    ws['H150'].number_format = '0.00%'
     
        # Por Ciudad en I: porcentaje por CO en una sola columna
    ws.cell(row=1, column=9, value='% Ciudad')
    co_totals = res_df.groupby('CO')['Cant'].sum().to_dict()
    for idx, row in res_df.iterrows():
        r = idx + 2
        total = co_totals.get(row['CO'], 1)
        ws.cell(row=r, column=9, value=f'=G{r}/{total}').number_format = '0.00%'

         # Por UN en J: aplicar fórmula condicional según UN
    ws.cell(row=1, column=10, value='% Comercio')
    commerce_codes = ['100','200','300','500','700']
    commerce_total = res_df.loc[res_df['UN'].isin(commerce_codes), 'Cant'].sum() or 1
    for idx, row in res_df.iterrows():
        r = idx + 2
        if row['UN'] in commerce_codes:
            ws.cell(row=r, column=10, value=f'=G{r}/{commerce_total}').number_format = '0.00%'
        else:
            ws.cell(row=r, column=10, value=f'=0.00%').number_format = '0.00%'
        
        # Por UN en K para cartera propia 500
    ws.cell(row=1, column=11, value='% Cartera Propia')
    propia_codes = ['500']
    propia_total = res_df.loc[res_df['UN'].isin(propia_codes), 'Cant'].sum() or 1
    for idx, row in res_df.iterrows():
        r=idx + 2
        if row['UN'] in propia_codes:
            ws.cell(row=r, column=11, value=f'=G{r}/{propia_total}').number_format = '0.00%'
        else:
            ws.cell(row=r, column=11, value=f'=0.00%').number_format = '0.00%'
            
             # Concatenación de D, E y F en L
    ws.cell(row=1, column=12, value='Concatenación')
    for idx, row in res_df.iterrows():
        r = idx + 2
        ws.cell(row=r, column=12, value=f'=D{r}&E{r}&F{r}')
           

print('✅ Total Resultado:', res_df['Cant'].sum())
print('✅ Total Resumen:', df_original['CEDULA'].nunique())

# -----------------------------
# 2) Generar mini tabla UGP
# -----------------------------
file_rent = "2025 Información Rentabilidad Carteras - Analítica.xlsx"
# Mes anterior en minúscula
mes_map2 = {
    'January':'enero','February':'febrero','March':'marzo','April':'abril',
    'May':'mayo','June':'junio','July':'julio','August':'agosto',
    'September':'septiembre','October':'octubre','November':'noviembre','December':'diciembre'
}
last_date = datetime.now().replace(day=1) - timedelta(days=1)
mp = mes_map2[last_date.strftime('%B')]
mp2 = mes_map2[(last_date.replace(day=1) - timedelta(days=1)).strftime('%B')]

# Leer hoja UGP del archivo de rentabilidad
ugp_wb = load_workbook(file_rent, data_only=True)
df_ugp = pd.read_excel(file_rent, sheet_name='UGP', header=2, dtype=str)
df_ugp.columns = df_ugp.columns.str.strip()

# Columnas: LINEA DE NEGOCIO, MES y Totales
line_col, m_col, *_, t_col = df_ugp.columns.tolist()
line_col = 'LINEA DE NEGOCIO'
m_col = 'MES'
t_col = '# Totales'
# Normalizar mes a minúscula
df_ugp[m_col] = df_ugp[m_col].str.lower()
# Construir mini resumen
mini = []

# Agrupar y sumar los # Totales por cada mes y cada línea de negocio
df_ugp[t_col] = pd.to_numeric(df_ugp[t_col], errors='coerce').fillna(0)

resumen = (
    df_ugp
    .groupby(['LINEA DE NEGOCIO', 'MES'], as_index=False)['# Totales']
    .sum()
    .sort_values(['LINEA DE NEGOCIO', 'MES'])
)

# Imprimir el resultado en consola
print("\n# Totales por cada mes y línea de negocio:")
print(resumen)

# Si quieres guardar este resumen en Excel:
with pd.ExcelWriter('Resultado_distribucion.xlsx', engine='openpyxl', mode='a') as writer:
    resumen.to_excel(writer, index=False, sheet_name='UGP_MES')

# Obtener el mes anterior en formato fecha igual al de tu columna MES
mes_anterior = (datetime.now().replace(day=1) - timedelta(days=1)).strftime('%Y-%m-%d')

# Filtrar solo los registros del mes anterior
df_mes_ant = df_ugp[df_ugp[m_col].str.contains(mes_anterior)]

# Agrupar y sumar los # Totales por cada línea de negocio del mes anterior
df_mes_ant[t_col] = pd.to_numeric(df_mes_ant[t_col], errors='coerce').fillna(0)
mini_ugp = (
    df_mes_ant
    .groupby(line_col, as_index=False)[t_col]
    .sum()
    .rename(columns={line_col: 'Categoria', t_col: 'Clientes'})
)
# ...existing code...

# Ordenar para que 'Total UGP' siempre esté al final
if 'Total UGP' in mini_ugp['Categoria'].values:
    mini_ugp_total = mini_ugp[mini_ugp['Categoria'] == 'Total UGP']
    mini_ugp_otros = mini_ugp[mini_ugp['Categoria'] != 'Total UGP']
    mini_ugp = pd.concat([mini_ugp_otros, mini_ugp_total], ignore_index=True)

# Imprimir resultado en consola
print("\nMini tabla UGP solo mes anterior:")
print(mini_ugp)


# Calcular columna 'UGP %' como valor directamente (excluyendo 'Total UGP')
mini_ugp = mini_ugp[~mini_ugp['Categoria'].str.contains("Total UGP", case=False, na=False)]
total_clientes = mini_ugp['Clientes'].sum()
mini_ugp['UGP %'] = (mini_ugp['Clientes'] / total_clientes).round(2)

# Agregar fila "Total UGP" con suma y 100%
fila_total = pd.DataFrame([{
    "Categoria": "Total UGP",
    "Clientes": total_clientes,
    "UGP %": 1.0
}])
mini_ugp = pd.concat([mini_ugp, fila_total], ignore_index=True)

    # =====================
    # 3) Resumen UN=600 por CO en UGP_MES_ANTERIOR
    # =====================
with pd.ExcelWriter('Resultado_distribucion.xlsx',
                    engine='openpyxl',
                    mode='a',
                    if_sheet_exists='replace') as writer:
    mini_ugp.to_excel(writer,
                      index=False,
                      sheet_name='UGP_MES_ANTERIOR')


# 3) Distribución proporcional en UGP_MES_ANTERIOR
file_path = 'Resultado_distribucion.xlsx'
wb = load_workbook(file_path)
ws = wb['UGP_MES_ANTERIOR']

# Formateo % en columna C
for i in range(2, 2 + len(mini_ugp)):
    ws.cell(row=i, column=3).number_format = '0.00%'

# Encabezados F–J
headers = ['Key','Cant','Vigente','Juridico','Castigado']
for col, h in enumerate(headers, start=6):
    ws.cell(row=1, column=col, value=h)

# Leemos porcentajes de C2–C4
pct = {
    'Vigente':   ws.cell(row=4, column=3).value or 0,
    'Juridico':  ws.cell(row=3, column=3).value or 0,
    'Castigado': ws.cell(row=2, column=3).value or 0
}

# Distribuimos por cada CO
for idx, co in enumerate(sorted(res_df['CO'].unique()), start=2):
    total = int(res_df.loc[(res_df['CO']==co)&(res_df['UN']=='600'),'Cant'].sum())
    key   = f"{co}600660104"
    alloc = { t: int(total * pct[t]) for t in pct }
    rem   = total - sum(alloc.values())
    if rem > 0:
        top = max(pct, key=pct.get)
        alloc[top] += rem

    ws.cell(row=idx, column=6, value=key)
    ws.cell(row=idx, column=7, value=total)
    ws.cell(row=idx, column=8, value=alloc['Vigente'])
    ws.cell(row=idx, column=9, value=alloc['Juridico'])
    ws.cell(row=idx, column=10,value=alloc['Castigado'])

wb.save(file_path)
   
   

# ✅ 1. Abrir archivo con data_only=True para obtener valores (no fórmulas)
wb = load_workbook(file_path, data_only=True)
ws = wb["UGP_MES_ANTERIOR"]

# ✅ 2. Leer tabla de Key, Cant, Vigente, Jurídico, Castigado (columnas F:J)
manual_table = []
for row in ws.iter_rows(min_row=2, max_row=150, min_col=6, max_col=10, values_only=True):
    key, cant, vigente, juridico, castigado = row
    if key:
        manual_table.append({
            "Concat": str(key),
            "CO": str(key)[:3],
            "Vigente": int(vigente or 0),
            "Juridico": int(juridico or 0),
            "Castigado": int(castigado or 0)
        })

df_manual = pd.DataFrame(manual_table)

# ✅ 3. Crear filas divididas según tipo
map_un_ccos = {
    "Vigente":   ("100", "110101"),
    "Juridico":  ("200", "220101"),
    "Castigado": ("300", "330101")
}

filas_tabla = []
for _, row in df_manual.iterrows():
    for tipo, (un, ccos) in map_un_ccos.items():
        cantidad = row[tipo]
        if cantidad > 0:
            filas_tabla.append({
                "Concat": row["Concat"],
                "UGP": "020",
                "CO": row["CO"],
                "UN": un,
                "CCOS": ccos,
                "Cant": cantidad
            })

df_tabla = pd.DataFrame(filas_tabla)

# ✅ 4. Guardar como hoja nueva 'tabla'
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_tabla.to_excel(writer, sheet_name='tabla', index=False)

print("✅ Hoja 'tabla' creada correctamente solo con valores (sin fórmulas).")

def construir_hoja_tabla(file_path: str):

    
    # -----------------------------
    # 1) Leemos Resultado (para traer todo excepto UN=600)
    # -----------------------------
    df_res = pd.read_excel(file_path, sheet_name="Resultado", dtype=str)
    df_res['Cant'] = pd.to_numeric(df_res['Cant'], errors='coerce').fillna(0).astype(int)
    df_res['Concat'] = df_res['CO'] + df_res['UN'] + df_res['CCOS']

    # Filas que NO son 600 -> van tal cual, con UGP = CO
    df_otros = df_res[df_res['UN'] != '600'].copy()
    df_otros['UGP'] = df_otros['CO']

    df_otros = df_otros[['Concat', 'UGP', 'CO', 'UN', 'CCOS', 'Cant']]

    # -----------------------------
    # 2) Leemos la distribución desde UGP_MES_ANTERIOR (valores, no fórmulas)
    # -----------------------------
    wb = load_workbook(file_path, data_only=True)
    ws = wb['UGP_MES_ANTERIOR']

    dist_rows = []
    # Key, Cant, Vigente, Juridico, Castigado están en F:J
    for row in ws.iter_rows(min_row=2, max_row=300, min_col=6, max_col=10, values_only=True):
        key, cant, vigente, juridico, castigado = row
        if key is None:
            continue
        dist_rows.append({
            'Concat': str(key),
            'CO': str(key)[:3],
            'Vigente': float(vigente or 0),
            'Juridico': float(juridico or 0),
            'Castigado': float(castigado or 0)
        })
    df_dist = pd.DataFrame(dist_rows)

    # -----------------------------
    # 3) Expandimos la UN=600 en 3 filas (100/200/300) con CCOS correspondientes
    # -----------------------------
    map_un_ccos = {
        'Vigente':   ('100', '110101'),
        'Juridico':  ('200', '220101'),
        'Castigado': ('300', '330101'),
    }

    nuevas_filas = []
    for _, r in df_dist.iterrows():
        for col, (un, ccos) in map_un_ccos.items():
            cant = r[col]
            if cant > 0 or True:  # si quieres también filas con 0, deja "or True"; si no, quítalo
                nuevas_filas.append({
                    'Concat': r['Concat'],
                    'UGP': '020',        # fijo para todo lo que viene de la 600
                    'CO': r['CO'],
                    'UN': un,
                    'CCOS': ccos,
                    'Cant': cant
                })
    df_ugp_expandida = pd.DataFrame(nuevas_filas, columns=['Concat','UGP','CO','UN','CCOS','Cant'])

    # -----------------------------
    # 4) Concatenamos: otros (sin 600) + lo expandido (600 repartida)
    # -----------------------------
    df_tabla = pd.concat([df_otros, df_ugp_expandida], ignore_index=True)

    # -----------------------------
    # 5) Guardamos en la hoja "tabla" (reemplazándola si existe)
    # -----------------------------
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_tabla.to_excel(writer, sheet_name='tabla', index=False)

    print("✅ Hoja 'tabla' creada/actualizada correctamente.")
    return df_tabla

# Ejecuta
file_path = "Resultado_distribucion.xlsx"
df_tabla = construir_hoja_tabla(file_path)
'''
# Función para quitar duplicados en la hoja tabla
def eliminar_duplicados_tabla(file_path):
    print("🔧 Eliminando duplicados por Concat...")
    
    # Leer hoja actual
    df_tabla = pd.read_excel(file_path, sheet_name='tabla', dtype=str).fillna('')
    
    print(f"📊 Filas antes de limpiar: {len(df_tabla)}")
    
    # Normalizar campos y convertir cantidad a numérico
    df_tabla['CO'] = df_tabla['CO'].str.zfill(3)
    df_tabla['UN'] = df_tabla['UN'].str.zfill(3)
    df_tabla['CCOS'] = df_tabla['CCOS'].str.zfill(6)
    df_tabla['Cant'] = pd.to_numeric(df_tabla['Cant'], errors='coerce').fillna(0)
    
    # Si no existe la columna Concat, la creamos
    if 'Concat' not in df_tabla.columns:
        df_tabla['Concat'] = df_tabla['CO'] + df_tabla['UN'] + df_tabla['CCOS']
    
    # Mostrar duplicados antes de agrupar
    duplicados = df_tabla[df_tabla.duplicated(subset=['Concat'], keep=False)]
    if len(duplicados) > 0:
        print(f"🔍 Se encontraron {len(duplicados)} registros duplicados:")
        print(duplicados[['Concat', 'UGP', 'CO', 'UN', 'CCOS', 'Cant']].to_string(index=False))
    
    # Agrupar por Concat y sumar cantidades
    df_tabla_clean = (
        df_tabla
        .groupby('Concat', as_index=False)
        .agg({
            'UGP': 'first',
            'CO': 'first',
            'UN': 'first',
            'CCOS': 'first',
            'Cant': 'sum'
        })
    )
    
    # Reordenar columnas
    df_tabla_clean = df_tabla_clean[['Concat', 'UGP', 'CO', 'UN', 'CCOS', 'Cant']]
    
    print(f"📊 Filas después de limpiar: {len(df_tabla_clean)}")
    print(f"✂️ Se eliminaron {len(df_tabla) - len(df_tabla_clean)} filas duplicadas")
    
    # Guardar hoja limpia
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_tabla_clean.to_excel(writer, sheet_name='tabla', index=False)
    
    print(f"✅ Consolidación completa: {len(df_tabla_clean)} filas únicas.")
    return df_tabla_clean

# Ejecutar eliminación de duplicados
df_tabla_clean = eliminar_duplicados_tabla(file_path)
'''

# Y usa directamente el DataFrame leído:
# Consolidar y eliminar duplicados por UGP, CO, UN, CCOS, sumando Cant
df_tabla = pd.read_excel(file_path, sheet_name='tabla', dtype=str).fillna('')
df_tabla['Cant'] = pd.to_numeric(df_tabla['Cant'], errors='coerce').fillna(0)

cols_to_keep = [col for col in df_tabla.columns if col not in ['Cant']]
df_tabla_grouped = (
    df_tabla
    .groupby(['UGP', 'CO', 'UN', 'CCOS'], as_index=False)
    .agg({**{col: 'first' for col in cols_to_keep}, 'Cant': 'sum'})
)

# Reordenar columnas si lo deseas
df_tabla_grouped = df_tabla_grouped[df_tabla.columns]

# Guardar el resultado en la hoja 'tabla'
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_tabla_grouped.to_excel(writer, sheet_name='tabla', index=False)

print(f"✅ Filas duplicadas eliminadas y cantidades sumadas por UGP, CO, UN, CCOS.")

#########################################
# Abre el archivo y edita la hoja 'tabla' con los datos limpios

with pd.ExcelWriter('Resultado_distribucion.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_tabla_grouped.to_excel(writer, index=False, sheet_name='tabla')
    wb = writer.book
    ws = writer.sheets['tabla']

    # Encabezados
    ws.cell(row=1, column=7, value='% Total')
    ws.cell(row=1, column=8, value='% Ciudad')
    ws.cell(row=1, column=9, value='% Comercio')
    ws.cell(row=1, column=10, value='% Cartera Propia')
    ws.cell(row=1, column=11, value='% UGP')


# 🔵 Total general
total_general = df_tabla_grouped['Cant'].sum()
ws['F173'] = total_general
ws['F173'].fill = PatternFill(fill_type='solid', fgColor='0000FF')

# ✅ % Total
for r in range(2, len(df_tabla_grouped) + 2):
    cant = df_tabla_grouped.iloc[r - 2]['Cant']
    ws.cell(row=r, column=7, value=cant / total_general if total_general else 0).number_format = '0.00%'

# Suma de % Total
ws[f'G{len(df_tabla_grouped) + 2}'] = sum([ws.cell(row=r, column=7).value for r in range(2, len(df_tabla_grouped) + 2)])
ws[f'G{len(df_tabla_grouped) + 2}'].number_format = '0.00%'

# ✅ % Ciudad
co_totals = df_tabla_grouped.groupby('CO')['Cant'].sum().to_dict()
for idx, row in df_tabla_grouped.iterrows():
    r = idx + 2
    total = co_totals.get(row['CO'], 1)
    ws.cell(row=r, column=8, value=row['Cant'] / total if total else 0).number_format = '0.00%'

# ✅ % Comercio
comercio_codes = ['100', '200', '300', '500', '700']
total_comercio = df_tabla_grouped[df_tabla_grouped['UN'].isin(comercio_codes)]['Cant'].sum() or 1
for idx, row in df_tabla_grouped.iterrows():
    r = idx + 2
    valor = row['Cant'] / total_comercio if row['UN'] in comercio_codes else 0
    ws.cell(row=r, column=9, value=valor).number_format = '0.00%'

# ✅ % Cartera Propia
total_propia = df_tabla_grouped[df_tabla_grouped['UN'] == '500']['Cant'].sum() or 1
for idx, row in df_tabla_grouped.iterrows():
    r = idx + 2
    valor = row['Cant'] / total_propia if row['UN'] == '500' else 0
    ws.cell(row=r, column=10, value=valor).number_format = '0.00%'

# ✅ % UGP — solo si UGP = '020'
total_ugp = df_tabla_grouped[df_tabla_grouped['UGP'] == '020']['Cant'].sum() or 1
for idx, row in df_tabla_grouped.iterrows():
    r = idx + 2
    valor = row['Cant'] / total_ugp if row['UGP'] == '020' else 0
    ws.cell(row=r, column=11, value=valor).number_format = '0.00%'

# Guardar los cambios
wb.save(file_path)



file_path = 'Resultado_distribucion.xlsx'
wb = load_workbook(file_path)
print("Hojas en el libro:", wb.sheetnames)


# ================================
# 4) Agregar hoja SQL al archivo
# ================================


# Parámetros SQL
SERVER     = os.getenv('SQL_SERVER')
PORT       = os.getenv('SQL_PORT')
USER       = os.getenv('SQL_USER')
PASSWORD   = os.getenv('SQL_PASSWORD')
DATABASE   = os.getenv('SQL_DATABASE')
SCHEMA     = os.getenv('SQL_SCHEMA')
TABLE      = 't351_co_mov_docto'
TABLE1     = 't253_co_auxiliares'
TABLE2     = 't284_co_ccosto'
TABLE3     = 't285_co_centro_op'

# Fecha actual
now = datetime.now()
year_filter = now.year
month_filter = now.month

# Conexión SQL
conn_str = f"mssql+pymssql://{USER}:{PASSWORD}@{SERVER}:{PORT}/{DATABASE}"
engine = create_engine(conn_str)

# Consulta SQL
query = f"""
SELECT
  t351.f351_id_cia,
  t253.f253_id            AS auxiliar_codigo,
  t351.f351_id_co_mov     AS co_codigo,
  t351.f351_id_un,
  t284.f284_id            AS ccosto_codigo
FROM {SCHEMA}.{TABLE} AS t351
LEFT JOIN {SCHEMA}.{TABLE1} AS t253
  ON t351.f351_rowid_auxiliar = t253.f253_rowid
LEFT JOIN {SCHEMA}.{TABLE2} AS t284
  ON t351.f351_rowid_ccosto   = t284.f284_rowid
LEFT JOIN {SCHEMA}.{TABLE3} AS t285
  ON t351.f351_id_co_mov      = t285.f285_id
WHERE
  t351.f351_id_cia         = 1
  AND YEAR(t351.f351_fecha) = {year_filter}
  AND MONTH(t351.f351_fecha)= {month_filter}
  AND t351.f351_id_un       = 600
  AND t351.f351_ind_estado  = 1
  AND t253.f253_id LIKE '5%'
"""

print("📥 Ejecutando SQL...")
df_sql = pd.read_sql(query, engine)
engine.dispose()

# Ordenar columnas manualmente por seguridad (aunque ya salen así por el SELECT)
ordered_columns = [
    "f351_id_cia",
    "auxiliar_codigo",
    "co_codigo",
    "f351_id_un",
    "ccosto_codigo"
]
df_sql = df_sql[ordered_columns]

# Guardar en Excel
with pd.ExcelWriter("Resultado_distribucion.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_sql.to_excel(writer, index=False, sheet_name="SQL_UN600")

print("✅ Hoja SQL_UN600 agregada a Resultado_distribucion.xlsx con columnas reordenadas.")

# ================================
# 5) Crear hoja 'ciudad' con todas las combinaciones posibles para cada auxiliar
# ================================


def crear_hoja_ciudad(file_path):
    # Leer hoja SQL_UN600
    df_sql = pd.read_excel(file_path, sheet_name="SQL_UN600", dtype=str)
    df_sql = df_sql.fillna("")
    auxiliares = df_sql["auxiliar_codigo"].drop_duplicates().tolist()
    # Definir las combinaciones de débito exactamente como el ejemplo
    combinaciones_debito = [
        ("base", "100", "110101"),
        ("base", "200", "220101"),
        ("base", "200", "220102"),
        ("base", "300", "330101"),
        ("base", "500", "550101"),
        ("base", "700", "770101"),
        ("base", "999", "990101"),
        ("base", "999", "990102"),
        ("base", "999", "990103"),
        ("base", "999", "990201"),
        ("base", "999", "990202"),
        ("base", "999", "990203"),
        ("base", "999", "990204"),
        ("base", "999", "990301"),
        ("020", "100", "110101"),
        ("020", "200", "220101"),
        ("020", "300", "330101"),
    ]
    # Columnas de salida
    columnas = [
        "Auxiliar", "Centro de operación base", "Unidad de negocio base", "Auxiliar de centro de costos base",
        "Auxiliar", "Centro de operación debito", "Unidad de negocio debito", "Auxiliar de centro de costos debito",
        "Auxiliar", "Centro de operación credito", "Unidad de negocio credito", "Auxiliar de centro de costos credito"
    ]
    filas = []
    # COs de 001 a 019 (incluyendo ceros a la izquierda), excluyendo 014 y 017
    cos = [str(i).zfill(3) for i in list(range(1, 20)) if i not in (14, 17)]
    for aux in auxiliares:
        for co in cos:
            un = "600"
            ccos_base = "660101"  # Fijo para ciudad
            for co_deb, un_deb, ccos_deb in combinaciones_debito:
                # Solo permitir combinaciones con UN_debito == 999 si co == '001'
                if un_deb == "999" and co != "001":
                    continue
                co_debito = co if co_deb == "base" else co_deb
                filas.append([
                    aux, co, un, ccos_base,
                    aux, co_debito, un_deb, ccos_deb,
                    aux, co, un, ccos_base
                ])
    df_ciudad = pd.DataFrame(filas, columns=columnas)

    # Leer hoja 'tabla' y obtener % Ciudad
    hoja_tabla = 'tabla'
    df_tabla = pd.read_excel(file_path, sheet_name=hoja_tabla, dtype=str).fillna('')
    if '% Ciudad' in df_tabla.columns:
        df_tabla['% Ciudad'] = pd.to_numeric(df_tabla['% Ciudad'], errors='coerce')
    else:
        # Si no existe la columna, poner 0
        df_tabla['% Ciudad'] = 0.0

    # Renombrar columnas para merge
    df_merge = df_tabla[['UGP', 'CO', 'UN', 'CCOS', '% Ciudad']].copy()

    # Realizar merge según el mapeo solicitado
    df_ciudad = df_ciudad.merge(
        df_merge,
        left_on=[
            'Centro de operación base',
            'Centro de operación debito',
            'Unidad de negocio debito',
            'Auxiliar de centro de costos debito'
        ],
        right_on=['CO', 'UGP', 'UN', 'CCOS'],
        how='left'
    )
    # Limpiar columnas extra
    df_ciudad.drop(columns=['UGP', 'CO', 'UN', 'CCOS'], inplace=True)
    # Rellenar y convertir a float
    df_ciudad['% Ciudad'] = df_ciudad['% Ciudad'].fillna(0.0).astype(float)


    
 
    wb = load_workbook(file_path)
    hoja_ciudad = 'ciudad'
    if hoja_ciudad in wb.sheetnames:
        del wb[hoja_ciudad]
    ws = wb.create_sheet(hoja_ciudad)
    for r_idx, row in enumerate(dataframe_to_rows(df_ciudad, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Formatear % Ciudad como porcentaje con dos decimales
            if df_ciudad.columns[c_idx - 1] == '% Ciudad' and r_idx > 1:
                cell.number_format = '0.00%'
    wb.save(file_path)
    print(f"✅ Hoja 'ciudad' creada con {len(df_ciudad)} filas y columna % Ciudad.")


# ================================
# Crear hoja 'Comercio' con ccos_base=660102
# ================================
def crear_hoja_comercio(file_path):
    # Leer hoja SQL_UN600
    df_sql = pd.read_excel(file_path, sheet_name="SQL_UN600", dtype=str)
    df_sql = df_sql.fillna("")
    auxiliares = df_sql["auxiliar_codigo"].drop_duplicates().tolist()
    # Definir las combinaciones de débito exactamente como el ejemplo
    combinaciones_debito = [
        ("base", "100", "110101"),
        ("base", "200", "220101"),
        ("base", "200", "220102"),
        ("base", "300", "330101"),
        ("base", "500", "550101"),
        ("base", "700", "770101"),
        ("base", "999", "990101"),
        ("base", "999", "990102"),
        ("base", "999", "990103"),
        ("base", "999", "990201"),
        ("base", "999", "990202"),
        ("base", "999", "990203"),
        ("base", "999", "990204"),
        ("base", "999", "990301"),
        ("020", "100", "110101"),
        ("020", "200", "220101"),
        ("020", "300", "330101"),
    ]
    columnas = [
        "Auxiliar", "Centro de operación base", "Unidad de negocio base", "Auxiliar de centro de costos base",
        "Auxiliar", "Centro de operación debito", "Unidad de negocio debito", "Auxiliar de centro de costos debito",
        "Auxiliar", "Centro de operación credito", "Unidad de negocio credito", "Auxiliar de centro de costos credito"
    ]
    filas = []
    # Solo CO base 001
    co_base = "001"
    un = "600"
    ccos_base = "660102"  # Fijo para comercio
    # COs de 001 a 019 (incluyendo ceros a la izquierda), excluyendo 014 y 017
    cos = [str(i).zfill(3) for i in list(range(1, 20)) if i not in (14, 17)]
    for aux in auxiliares:
        # Para cada combinación de débito, para cada CO destino
        for co_debito in cos + ["020"]:
            for un_deb, ccos_deb in [
                ("100", "110101"),
                ("200", "220101"),
                ("200", "220102"),
                ("300", "330101"),
                ("500", "550101"),
                ("700", "770101"),
                ("999", "990101"),
                ("999", "990102"),
                ("999", "990103"),
                ("999", "990201"),
                ("999", "990202"),
                ("999", "990203"),
                ("999", "990204"),
                ("999", "990301"),
            ]:
                # Solo permitir combinaciones con UN_debito == 999 si co_debito == '001'
                if un_deb == "999" and co_debito != "001":
                    continue
                filas.append([
                    aux, co_base, un, ccos_base,
                    aux, co_debito, un_deb, ccos_deb,
                    aux, co_base, un, ccos_base
                ])
        # Para CO debito = 020, solo las combinaciones especiales
        for un_deb, ccos_deb in [
            ("100", "110101"),
            ("200", "220101"),
            ("300", "330101")
        ]:
            filas.append([
                aux, co_base, un, ccos_base,
                aux, "020", un_deb, ccos_deb,
                aux, co_base, un, ccos_base
            ])
    df_comercio = pd.DataFrame(filas, columns=columnas)
    # Guardar en hoja 'Comercio'
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_comercio.to_excel(writer, index=False, sheet_name="Comercio")
    print(f"✅ Hoja 'Comercio' creada con {len(df_comercio)} filas.")

    def actualizar_porcentaje_comercio(file_path):
        # Leer hoja 'tabla' y asegurar que % Comercio es numérico
        df_tabla = pd.read_excel(file_path, sheet_name='tabla', dtype=str).fillna('')
        if '% Comercio' in df_tabla.columns:
            df_tabla['% Comercio'] = pd.to_numeric(df_tabla['% Comercio'], errors='coerce')
        else:
            df_tabla['% Comercio'] = 0.0

        # Renombrar columnas para merge
        df_merge1 = df_tabla[['UGP', 'CO', 'UN', 'CCOS', '% Comercio']].copy()
        df_merge2 = df_tabla[['CO', 'UGP', 'UN', 'CCOS', '% Comercio']].copy()

        # Leer hoja 'Comercio'
        df_comercio = pd.read_excel(file_path, sheet_name='Comercio', dtype=str).fillna('')

        # Primera llave: Centro de operación debito != '020'
        mask1 = df_comercio['Centro de operación debito'] != '020'
        df_comercio1 = df_comercio[mask1].merge(
            df_merge1,
            left_on=['Centro de operación debito', 'Centro de operación debito', 'Unidad de negocio debito', 'Auxiliar de centro de costos debito'],
            right_on=['UGP', 'CO', 'UN', 'CCOS'],
            how='left'
        )

        # Segunda llave: Centro de operación debito == '020'
        mask2 = df_comercio['Centro de operación debito'] == '020'
        df_comercio2 = df_comercio[mask2].merge(
            df_merge2,
            left_on=['Centro de operación base', 'Centro de operación debito', 'Unidad de negocio debito', 'Auxiliar de centro de costos debito'],
            right_on=['CO', 'UGP', 'UN', 'CCOS'],
            how='left'
        )

        # Unir ambos resultados
        df_comercio_actualizado = pd.concat([df_comercio1, df_comercio2], ignore_index=True)

        # Actualizar la columna % Comercio
        df_comercio_actualizado['% Comercio'] = df_comercio_actualizado['% Comercio'].fillna(0.0).astype(float)

        # Limpiar columnas extra del merge
        for col in ['UGP', 'CO', 'UN', 'CCOS']:
            if col in df_comercio_actualizado.columns:
                df_comercio_actualizado.drop(columns=col, inplace=True)

        # Guardar en la hoja 'Comercio' y formatear la columna % Comercio como porcentaje
        from openpyxl.utils.dataframe import dataframe_to_rows
        wb = load_workbook(file_path)
        if 'Comercio' in wb.sheetnames:
            del wb['Comercio']
        ws = wb.create_sheet('Comercio')
        for r_idx, row in enumerate(dataframe_to_rows(df_comercio_actualizado, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Formatear % Comercio como porcentaje con dos decimales
                if df_comercio_actualizado.columns[c_idx - 1] == '% Comercio' and r_idx > 1:
                    cell.number_format = '0.00%'
        wb.save(file_path)
        print(f"✅ Hoja 'Comercio' actualizada con columna % Comercio.")


    # Ejecuta la función
    actualizar_porcentaje_comercio("Resultado_distribucion.xlsx")
# Ejecutar la función para crear la hoja ciudad
crear_hoja_ciudad("Resultado_distribucion.xlsx")

# Ejecutar la función para crear la hoja Comercio
crear_hoja_comercio("Resultado_distribucion.xlsx")


